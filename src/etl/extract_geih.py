"""
Extracción de datos de la GEIH del DANE.

Este módulo incluye dos estrategias:
1. `generate_realistic_data()`: Genera datos sintéticos realistas a nivel departamental
   basados en patrones históricos conocidos del desempleo en Colombia.
   Útil cuando los anexos Excel del DANE no están disponibles o no tienen
   desglose departamental.
2. `extract_geih_excel()`: Parsea los anexos Excel reales del DANE (work in progress).

Los datos sintéticos se calibraron con información pública del DANE para que
las tendencias y valores sean aproximados a la realidad.
"""

from pathlib import Path

import numpy as np
import pandas as pd
from loguru import logger

# ─── Datos base realistas de desempleo departamental ─────────────────────────
# Tasa de desempleo promedio 2015-2024 aproximada por departamento
# Fuente: DANE GEIH - valores de referencia públicos
DEPARTMENT_BASELINE = {
    "AMAZONAS": 11.5,
    "ANTIOQUIA": 9.8,
    "ARAUCA": 10.2,
    "ATLANTICO": 8.5,
    "BOLIVAR": 7.8,
    "BOYACA": 8.2,
    "CALDAS": 10.5,
    "CAQUETA": 11.8,
    "CASANARE": 9.5,
    "CAUCA": 11.2,
    "CESAR": 9.8,
    "CHOCO": 15.5,
    "CORDOBA": 10.8,
    "CUNDINAMARCA": 8.0,
    "GUAINIA": 12.0,
    "GUAVIARE": 10.5,
    "HUILA": 8.8,
    "LA GUAJIRA": 13.2,
    "MAGDALENA": 9.5,
    "META": 9.0,
    "NARINO": 10.0,
    "NORTE DE SANTANDER": 13.5,
    "PUTUMAYO": 10.8,
    "QUINDIO": 13.8,
    "RISARALDA": 9.5,
    "SAN ANDRES": 8.5,
    "SANTANDER": 7.5,
    "SUCRE": 9.2,
    "TOLIMA": 11.5,
    "VALLE DEL CAUCA": 10.8,
    "VAUPES": 11.0,
    "VICHADA": 10.5,
    "BOGOTA D.C.": 9.0,
}

# Variación estacional mensual (factor multiplicativo)
SEASONAL_FACTORS = {
    1: 1.15,   # Enero: alto (fin de contratos temporales navideños)
    2: 1.10,   # Febrero: alto
    3: 1.05,   # Marzo: medio-alto
    4: 0.95,   # Abril: medio-bajo
    5: 0.92,   # Mayo: bajo
    6: 1.00,   # Junio: medio (nuevos graduados)
    7: 1.08,   # Julio: medio-alto
    8: 1.02,   # Agosto: medio
    9: 0.98,   # Septiembre: medio-bajo
    10: 0.95,  # Octubre: bajo
    11: 0.90,  # Noviembre: bajo (contrataciones navideñas)
    12: 0.85,  # Diciembre: más bajo (empleo temporal navideño)
}

# Tendencia nacional: bajada gradual desde 2015 (~9.5%) hasta 2019 (~10.5%),
# pico COVID en 2020 (~16%), recuperación hasta 2026 (~8.8%)
YEAR_TREND = {
    2015: 0.0,
    2016: 0.3,
    2017: 0.1,
    2018: -0.2,
    2019: 0.5,
    2020: 6.0,    # COVID spike
    2021: 3.0,    # Recuperación parcial
    2022: 1.5,    # Recuperación
    2023: 0.5,    # Casi normalizado
    2024: 0.2,
    2025: -0.1,
    2026: -0.3,
}

DEPARTMENT_CODES = {
    "AMAZONAS": "91", "ANTIOQUIA": "05", "ARAUCA": "81", "ATLANTICO": "08",
    "BOLIVAR": "13", "BOYACA": "15", "CALDAS": "17", "CAQUETA": "18",
    "CASANARE": "85", "CAUCA": "19", "CESAR": "20", "CHOCO": "27",
    "CORDOBA": "23", "CUNDINAMARCA": "25", "GUAINIA": "94", "GUAVIARE": "95",
    "HUILA": "41", "LA GUAJIRA": "44", "MAGDALENA": "47", "META": "50",
    "NARINO": "52", "NORTE DE SANTANDER": "54", "PUTUMAYO": "86",
    "QUINDIO": "63", "RISARALDA": "66", "SAN ANDRES": "88",
    "SANTANDER": "68", "SUCRE": "70", "TOLIMA": "73",
    "VALLE DEL CAUCA": "76", "VAUPES": "97", "VICHADA": "99",
    "BOGOTA D.C.": "11",
}


def generate_realistic_data(
    start_year: int = 2015,
    end_year: int = 2026,
    seed: int = 42,
) -> pd.DataFrame:
    """
    Genera un DataFrame con datos sintéticos realistas de desempleo departamental.

    Los datos se basan en:
    - Tasas promedio conocidas por departamento (DANE GEIH)
    - Patrones estacionales mensuales
    - Tendencia nacional (COVID spike en 2020)
    - Variabilidad aleatoria controlada (ruido gaussiano)

    Args:
        start_year: Año inicial.
        end_year: Año final (inclusive).
        seed: Semilla aleatoria para reproducibilidad.

    Returns:
        DataFrame con columnas:
        - departamento, cod_departamento, año, mes,
        - tasa_desempleo, tasa_ocupacion, tgp
    """
    logger.info(f"Generando datos sintéticos {start_year}-{end_year}...")
    rng = np.random.default_rng(seed)

    records = []
    departamentos = sorted(DEPARTMENT_BASELINE.keys())

    for depto in departamentos:
        baseline = DEPARTMENT_BASELINE[depto]
        # Cada departamento tiene una desviación propia
        depto_volatility = rng.uniform(0.5, 2.0)

        for year in range(start_year, end_year + 1):
            year_trend = YEAR_TREND.get(year, 0.0)

            for month in range(1, 13):
                # No generar meses futuros
                if year == end_year and month > 4:  # Abril 2026 es el último dato disponible
                    continue
                seasonal = SEASONAL_FACTORS[month]
                noise = rng.normal(0, depto_volatility)

                # Tasa de desempleo con todos los factores
                td = baseline + year_trend + noise
                td = td * seasonal
                td = np.clip(td, 3.0, 35.0)  # Límites realistas

                # TGP (Tasa Global de Participación): inversamente correlacionada
                tgp_base = 64.0 - (baseline - 9.0) * 0.5  # ~62-67%
                tgp = tgp_base + rng.normal(0, 1.5)
                tgp = np.clip(tgp, 50.0, 80.0)

                # TO (Tasa de Ocupación) = TGP * (1 - TD/100)
                to = tgp * (1 - td / 100)
                to = np.clip(to, 35.0, 75.0)

                records.append({
                    "departamento": depto,
                    "cod_departamento": DEPARTMENT_CODES.get(depto, "00"),
                    "año": year,
                    "mes": month,
                    "tasa_desempleo": round(td, 1),
                    "tasa_ocupacion": round(to, 1),
                    "tgp": round(tgp, 1),
                })

    df = pd.DataFrame(records)
    logger.success(
        f"Datos generados: {len(df)} registros, {df['departamento'].nunique()} deptos, "
        f"{df['año'].min()}-{df['año'].max()}"
    )
    return df


def extract_geih_excel(filepath: Path) -> pd.DataFrame | None:
    """
    Parsea un archivo Excel de la GEIH del DANE.

    Actualmente extrae datos del Total Nacional (no departamental).
    Work in progress para datos departamentales.

    Args:
        filepath: Ruta al archivo .xlsx

    Returns:
        DataFrame con datos de la serie o None si no se puede parsear.
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)

        sheet_name = None
        for candidate in ["Total nacional", "Tnal mensual"]:
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break

        if sheet_name is None:
            logger.warning(f"No se encontró hoja de datos en {filepath.name}")
            wb.close()
            return None

        ws = wb[sheet_name]
        logger.info(f"Parseando {sheet_name} de {filepath.name}...")

        # Encontrar fila de inicio (donde dice "Concepto" y "2001")
        data_start = None
        for row in range(1, min(20, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "Concepto" in str(cell_val):
                data_start = row
                break

        if data_start is None:
            logger.warning("No se encontró inicio de datos")
            wb.close()
            return None

        # Meses de la fila data_start + 1 (ej: Ene, Feb, Mar...)
        months_row = data_start + 1
        months = []
        for col in range(2, ws.max_column + 1):
            month_val = ws.cell(row=months_row, column=col).value
            if month_val:
                months.append(str(month_val).strip()[:3])
            else:
                break

        logger.info(f"Meses encontrados: {len(months)} ({months[:3]}...{months[-3:]})")

        wb.close()
        # Por ahora retornamos None hasta implementar la extracción completa
        return None

    except ImportError:
        logger.error("openpyxl no instalado")
        return None
    except Exception as e:
        logger.error(f"Error parseando {filepath}: {e}")
        return None
